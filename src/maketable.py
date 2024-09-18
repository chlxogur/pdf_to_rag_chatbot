from src.pdfparsing import read_text_file, extract_table_with_won_unit, table_to_dic
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill
from openpyxl.styles import Alignment
from collections import deque
from src.config import OUTPUT_PATH

def make_column(df, table_name, worksheet, max_depth, target_year = False):
    company_dic = {}
    company_list = df.keys()
    rightmost_column_number = max_depth
    all_years = []
    for company in company_list:        # 여기부터 기록 시작
        if isinstance(df[company][table_name], pd.DataFrame):
            if company_dic.get(company) is None:                    
                company_dic[company] = {
                    "left":rightmost_column_number + 1
                }
                # 원하는 년도만 고르는 기능
                if target_year == False:
                    current_year = df[company][table_name].columns[1:]
                else:
                    current_year = [year for year in df[company][table_name].columns[1:] if year in target_year]
                    
                all_years.extend(current_year)
                for year_value in current_year:
                    company_dic[company].update({year_value:rightmost_column_number + 1})
                    cell = worksheet.cell(row = 2, column = rightmost_column_number + 1, value=year_value + "년")
                    cell.alignment = Alignment(horizontal="center")
                    rightmost_column_number += 1
                company_dic[company].update({"right":rightmost_column_number})
                worksheet.merge_cells(start_row=1, start_column=company_dic[company]["left"], end_row=1, end_column=rightmost_column_number)
                cell = worksheet.cell(row = 1, column = company_dic[company]["left"], value=company.capitalize())       # 회사명 입력
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill(fgColor = "C0C0C0", fill_type="solid")
    
    for company in ["sum", "average"]:          # 합계랑 평균
        if company_dic.get(company) is None:                    
                company_dic[company] = {
                    "left":rightmost_column_number + 1
                }
                all_years = sorted(list(set(all_years)), reverse=True)
                for year_value in all_years:
                    company_dic[company].update({year_value:rightmost_column_number + 1})
                    cell = worksheet.cell(row = 2, column = rightmost_column_number + 1, value=year_value + "년")
                    cell.alignment = Alignment(horizontal="center")
                    rightmost_column_number += 1
                company_dic[company].update({"right":rightmost_column_number})
                worksheet.merge_cells(start_row=1, start_column=company_dic[company]["left"], end_row=1, end_column=rightmost_column_number)
                cell = worksheet.cell(row = 1, column = company_dic[company]["left"], value=company.capitalize())       # 회사명 입력
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill(fgColor = "C0C0C0", fill_type="solid")
    return worksheet, company_dic

def make_table():
    CONTENT_START_ROW = 4           # 엑셀 형식에따라
    df = {}
    file_names, desired_table_name_list, target_year = read_text_file()
    for file_name in file_names:
        df[file_name] = extract_table_with_won_unit(file_name, desired_table_name_list)
        
    data = table_to_dic(df, target_year)
        
    for table_name in data.keys():
        workbook = Workbook()
        worksheet = workbook.active
        max_depth = 0

        for file_name in file_names:
            if isinstance(df[file_name][table_name], pd.DataFrame):
                depth = df[file_name][table_name]["과목"].apply(lambda x: x.count("_")).max()
                if max_depth < depth:
                    max_depth = depth
        max_depth += 1                  # _가 하나도 없으면 depth를 1로 정하려고.
        current_row = CONTENT_START_ROW
        worksheet, company_dic = make_column(df, table_name, worksheet, max_depth, target_year)
        dfsearch_stack = deque()
        dfsearch_stack.append((data[table_name], []))
        while dfsearch_stack:
            current_dic, ancestors = dfsearch_stack.pop()
            for key, value in reversed(current_dic.items()):
                current_path = ancestors + [key]
                if isinstance(value, dict) and key != "Value":
                    dfsearch_stack.append((value, current_path))
                else:
                    left_index = 1
                    for path_node in current_path[:-1]:                                 # 과목명을 다중컬럼으로 입력
                        cell = worksheet.cell(row = current_row, column=left_index, value = path_node)
                        left_index += 1
                    for key_of_company, value_of_company in value.items():
                        for key_of_cell, value_of_cell in value_of_company.items():
                            cell = worksheet.cell(row=current_row, column = company_dic[key_of_company][key_of_cell], value = value_of_cell)
                    current_row += 1
        
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row = 3, column = col)
            cell.fill = PatternFill(fgColor = "000000", fill_type="solid")
            
        for row in range(4, worksheet.max_row + 1):
            for col in range(max_depth + 1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                
                if cell.value is None or np.isnan(cell.value) or str(cell.value).strip() == '':
                    cell.value = '-'
                    cell.alignment = Alignment(horizontal="right")
                    
        # 외곽선 싹 긋기
        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border

        # 셀 너비 조정
        for column_cells in worksheet.columns:
            max_length = 0
            column = None

            for row_idx, cell in enumerate(column_cells, start=1):
                if row_idx == 1:
                    continue
                
                if column is None:  # 병합되지 않은 셀의 열 알파벳을 얻음
                    column = cell.column_letter
                    
                try:
                    # 셀의 값이 문자열이거나 숫자인 경우 길이를 계산
                    if isinstance(cell.value, str):
                        max_length = max(max_length, len(str(cell.value)) * 1.8)
                    elif cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            # 엑셀의 기본 글꼴 기준으로 열 너비 조정 (조정값을 곱해 최적화)
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_depth)
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_depth)
        workbook.save(OUTPUT_PATH + f"{table_name}.xlsx")    
